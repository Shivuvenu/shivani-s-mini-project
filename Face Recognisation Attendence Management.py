import cv2
import numpy as np
import os
from tkinter import *
from tkinter import messagebox
from PIL import Image
from datetime import datetime
import openpyxl
import json

# Paths
training_data_dir = 'training_data/'
model_path = 'trainer.yml'
attendance_file = 'attendance.xlsx'
person_ids_file = 'person_ids.json'  # New file to store person IDs

# Ensure the directories exist
if not os.path.exists(training_data_dir):
    os.makedirs(training_data_dir)

# Initialize OpenCV's face recognizer
recognizer = cv2.face.LBPHFaceRecognizer_create()

# Load OpenCV's pre-trained Haar Cascade for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Initialize Excel file for attendance if it doesn't exist
if not os.path.exists(attendance_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Attendance"
    sheet.append(["Name", "Date", "Time"])  # Add headers
    workbook.save(attendance_file)

def capture_faces(person_name):
    """Captures 100 images of the given person for training."""
    if not person_name:
        messagebox.showwarning("Warning", "Please enter a student name.")
        return

    cap = cv2.VideoCapture(0)
    count = 0

    person_path = os.path.join(training_data_dir, person_name)
    if not os.path.exists(person_path):
        os.makedirs(person_path)
    
    while count < 100:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Error", "Failed to capture image from camera.")
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            count += 1
            face = gray[y:y+h, x:x+w]
            face_resized = cv2.resize(face, (200, 200))
            cv2.imwrite(f"{person_path}/image_{count}.jpg", face_resized)
            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
            cv2.putText(frame, f"Captured {count}/100", (x-10, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0), 2)

        cv2.imshow("Capturing Faces", frame)
        if cv2.waitKey(1) == 27:  # Press 'ESC' to exit
            break
        if count >= 100:
            break

    cap.release()
    cv2.destroyAllWindows()
    if count == 100:
        messagebox.showinfo("Info", "100 Images Captured Successfully. Now you can train the model.")
    else:
        messagebox.showwarning("Warning", "Image capture was interrupted or incomplete.")

def train_model():
    """Trains the face recognizer model using collected images."""
    faces = []
    ids = []
    current_id = 0
    person_ids = {}

    # Get sorted list of person names to ensure consistent IDs
    person_names = sorted(os.listdir(training_data_dir))

    for person_name in person_names:
        person_path = os.path.join(training_data_dir, person_name)
        if not os.path.isdir(person_path):
            continue

        person_ids[person_name] = current_id
        current_id += 1

        for img_name in os.listdir(person_path):
            img_path = os.path.join(person_path, img_name)
            img = Image.open(img_path).convert('L')  # Convert to grayscale
            img_np = np.array(img, 'uint8')
            faces.append(img_np)
            ids.append(person_ids[person_name])  # Use numeric IDs for training

    if faces and ids:
        recognizer.train(faces, np.array(ids))
        recognizer.write(model_path)
        # Save the mapping of person names to IDs
        with open(person_ids_file, 'w') as f:
            json.dump(person_ids, f)
        messagebox.showinfo("Info", "Model Trained Successfully")
    else:
        messagebox.showwarning("Warning", "No faces found to train the model.")

def recognize_face():
    """Recognizes faces in real-time and marks attendance."""
    if not os.path.exists(model_path):
        messagebox.showerror("Error", "Model not found. Please train the model first.")
        return
    if not os.path.exists(person_ids_file):
        messagebox.showerror("Error", "Person IDs not found. Please train the model first.")
        return

    cap = cv2.VideoCapture(0)
    recognizer.read(model_path)
    attendance = set()  # To keep track of already logged names

    # Load the mapping of person IDs to names
    with open(person_ids_file, 'r') as f:
        person_ids = json.load(f)
    # Convert person_ids to have integer keys
    person_ids = {name: int(id_) for name, id_ in person_ids.items()}
    id_to_name = {id_: name for name, id_ in person_ids.items()}

    while True:
        ret, frame = cap.read()
        if not ret:
            messagebox.showerror("Error", "Failed to capture image from camera.")
            break
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        for (x, y, w, h) in faces:
            face = gray[y:y+h, x:x+w]
            face_resized = cv2.resize(face, (200, 200))

            id_, confidence = recognizer.predict(face_resized)
            if confidence < 50:  # Confidence threshold
                name = id_to_name.get(id_)
                if name and name not in attendance:
                    # Log the attendance if this person hasn't been logged already
                    attendance.add(name)
                    now = datetime.now()
                    date = now.strftime("%Y-%m-%d")
                    time_str = now.strftime("%H:%M:%S")
                    log_attendance(name, date, time_str)
                    cv2.putText(frame, f"{name} {confidence:.2f}", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                    messagebox.showinfo("Attendance", f"Attendance Registered for {name}")
                    cap.release()  # Close the camera
                    cv2.destroyAllWindows()  # Close the recognition window
                    return  # Exit after recognizing and logging attendance
            else:
                cv2.putText(frame, "Unknown", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

            cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)

        cv2.imshow("Recognizing Faces", frame)
        if cv2.waitKey(1) == 27:  # Press 'ESC' to exit
            break

    cap.release()
    cv2.destroyAllWindows()

def log_attendance(name, date, time_str):
    """Logs attendance details into an Excel file."""
    if not os.path.exists(attendance_file):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time"])  # Add headers
    else:
        workbook = openpyxl.load_workbook(attendance_file)
        sheet = workbook.active

    # Check if attendance has already been marked for this person today
    rows = sheet.iter_rows(values_only=True)
    for row in rows:
        if row[0] == name and row[1] == date:
            messagebox.showinfo("Info", f"Attendance already marked for {name} today.")
            return

    sheet.append([name, date, time_str])
    workbook.save(attendance_file)
    messagebox.showinfo("Info", f"Attendance Registered for {name}")

def manual_attendance():
    """Allows manual attendance registration."""
    name = entry_name.get()
    if name:
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        log_attendance(name, date, time_str)
    else:
        messagebox.showwarning("Warning", "Name cannot be empty")

# GUI Setup
root = Tk()
root.title("Facial Recognition Attendance System")
root.geometry("500x400")

# Heading
Label(root, text="Attendance System", font=("Helvetica", 16)).pack(pady=10)

# Manual Attendance
frame_manual = Frame(root)
frame_manual.pack(pady=20)
Label(frame_manual, text="Student Name:").pack(side=LEFT)
entry_name = Entry(frame_manual, width=30)
entry_name.pack(side=LEFT)
Button(frame_manual, text="Mark Attendance", command=manual_attendance).pack(side=LEFT, padx=10)

# Buttons
frame_buttons = Frame(root)
frame_buttons.pack(pady=20)

Button(frame_buttons, text="Capture Faces", command=lambda: capture_faces(entry_name.get())).pack(side=LEFT, padx=10)
Button(frame_buttons, text="Train Model", command=train_model).pack(side=LEFT, padx=10)
Button(frame_buttons, text="Recognize Faces", command=recognize_face).pack(side=LEFT, padx=10)

root.mainloop()
